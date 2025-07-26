import os
import json
import base64
import uuid
import io
from typing import Optional
from fastapi import FastAPI, Request, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
import httpx
from dotenv import load_dotenv
from starlette.middleware.sessions import SessionMiddleware
from auth_dev import SECRET_KEY, get_current_user, require_auth, login_google, auth_callback, logout, check_demo_limits, increment_usage, is_admin_user

# PDF and file processing imports
try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from PIL import Image
    from docx import Document
    import openpyxl
    IMAGE_AVAILABLE = True
except ImportError:
    IMAGE_AVAILABLE = False

load_dotenv('.env.dev')

app = FastAPI(title="Zen's LiteLLM Chat App - Developer Demo")

# Add session middleware
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY)

# Templates
templates = Jinja2Templates(directory="templates")

# Configuration
LITELLM_API_BASE = os.getenv("LITELLM_API_BASE", "https://your-litellm-proxy.com")
API_KEY = os.getenv("LITELLM_API_KEY")
VERSION = os.getenv("APP_VERSION", "3.0.0-dev")

print("🚀 Zen's LiteLLM Chat App (Developer Demo) starting...")
print(f"📦 Version: {VERSION}")
print(f"🔗 API Base: {LITELLM_API_BASE}")
print(f"📄 PDF Support: {'✅' if PDF_AVAILABLE else '❌'}")
print(f"🖼️ Image Support: {'✅' if IMAGE_AVAILABLE else '❌'}")
print(f"🎭 Demo Mode: Enabled")

# In-memory storage for chat sessions (replace with database in production)
chat_sessions = {}

# Simple response cache for faster repeated queries
response_cache = {}

def process_file_content(file_data: bytes, file_name: str) -> str:
    """Process different file types and extract text content"""
    try:
        file_extension = file_name.lower().split('.')[-1]
        
        # Text files
        if file_extension in ['txt', 'json', 'yaml', 'yml', 'md', 'csv', 'log']:
            try:
                return file_data.decode('utf-8')
            except UnicodeDecodeError:
                return f"[Binary text file: {file_name}]"
        
        # PDF files
        elif file_extension == 'pdf' and PDF_AVAILABLE:
            try:
                pdf_file = io.BytesIO(file_data)
                pdf_reader = PyPDF2.PdfReader(pdf_file)
                text_content = []
                for page_num, page in enumerate(pdf_reader.pages, 1):
                    text_content.append(f"--- Page {page_num} ---")
                    text_content.append(page.extract_text())
                return "\n".join(text_content)
            except Exception as e:
                return f"[PDF processing error: {str(e)}]"
        
        # Word documents
        elif file_extension == 'docx' and IMAGE_AVAILABLE:
            try:
                doc = Document(io.BytesIO(file_data))
                text_content = []
                for paragraph in doc.paragraphs:
                    if paragraph.text.strip():
                        text_content.append(paragraph.text)
                return "\n".join(text_content)
            except Exception as e:
                return f"[DOCX processing error: {str(e)}]"
        
        # Excel files
        elif file_extension in ['xlsx', 'xls'] and IMAGE_AVAILABLE:
            try:
                workbook = openpyxl.load_workbook(io.BytesIO(file_data))
                text_content = []
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    text_content.append(f"--- Sheet: {sheet_name} ---")
                    for row in sheet.iter_rows(values_only=True):
                        if any(cell is not None for cell in row):
                            text_content.append("\t".join(str(cell) if cell is not None else "" for cell in row))
                return "\n".join(text_content)
            except Exception as e:
                return f"[Excel processing error: {str(e)}]"
        
        # Image files
        elif file_extension in ['jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp'] and IMAGE_AVAILABLE:
            try:
                image = Image.open(io.BytesIO(file_data))
                image_info = f"""
Image Analysis:
- Format: {image.format}
- Size: {image.size[0]}x{image.size[1]} pixels
- Mode: {image.mode}
- File: {file_name}
"""
                return image_info
            except Exception as e:
                return f"[Image processing error: {str(e)}]"
        
        else:
            return f"[Unsupported file type: {file_extension}]"
            
    except Exception as e:
        return f"[File processing error: {str(e)}]"

def is_image_generation_model(model: str) -> bool:
    """Check if the model supports image generation"""
    # Only true image generation models
    image_models = [
        'dall-e-3', 'dall-e-2', 'dall-e',
        'midjourney', 'stable-diffusion',
        'sdxl', 'kandinsky', 'deepfloyd'
    ]
    return any(img_model in model.lower() for img_model in image_models)

def generate_image_prompt(user_message: str) -> str:
    """Extract or enhance image generation prompt from user message"""
    # Look for image generation keywords
    image_keywords = ['generate image', 'create image', 'draw', 'picture', 'photo', 'image of']
    
    if any(keyword in user_message.lower() for keyword in image_keywords):
        return user_message
    else:
        # Enhance the prompt for better image generation
        return f"Create a high-quality, detailed image of: {user_message}"

async def handle_image_generation(request: Request, message: str, model: str, session_id: str, file_content: Optional[str], file_name: Optional[str]):
    """Handle image generation requests"""
    user = require_auth(request)
    
    # Check demo limits for non-admin users
    can_proceed, error_message = check_demo_limits(user['email'])
    if not can_proceed:
        return {"error": error_message, "file_name": file_name}
    
    # Prepare the image generation prompt
    image_prompt = generate_image_prompt(message)
    
    # Add file context if provided
    if file_content and file_name:
        try:
            file_data = base64.b64decode(file_content)
            file_text = process_file_content(file_data, file_name)
            if file_text:
                image_prompt = f"{image_prompt}\n\nContext from uploaded file ({file_name}):\n{file_text[:1000]}"
        except Exception as e:
            print(f"Error processing file for image generation: {e}")
    
    # Prepare payload for image generation
    payload = {
        "model": model,
        "prompt": image_prompt,
        "n": 1,
        "size": "1024x1024",
        "quality": "standard",
        "response_format": "url"
    }
    
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {API_KEY}"
    }
    
    try:
        async with httpx.AsyncClient() as client:
            # Try image generation endpoint first
            response = await client.post(
                f"{LITELLM_API_BASE}/images/generations",
                json=payload,
                headers=headers,
                timeout=30.0
            )
            
            if response.status_code == 404:
                # Fallback to chat completions with image generation instruction
                chat_payload = {
                    "model": model,
                    "messages": [
                        {
                            "role": "system",
                            "content": "You are an AI assistant that can generate images. When asked to create an image, provide a detailed description and suggest using an image generation service."
                        },
                        {
                            "role": "user",
                            "content": f"Generate an image based on this description: {image_prompt}"
                        }
                    ],
                    "stream": False,
                    "max_tokens": 1000,
                    "temperature": 0.7
                }
                
                response = await client.post(
                    f"{LITELLM_API_BASE}/chat/completions",
                    json=chat_payload,
                    headers=headers,
                    timeout=15.0
                )
            
            response.raise_for_status()
            result = response.json()
            
            if "data" in result and result["data"]:
                # Image generation successful
                image_url = result["data"][0]["url"]
                assistant_message = f"🎨 **Image Generated Successfully!**\n\n### 🖼️ **Generated Image**\n![Generated Image]({image_url})\n\n### 📝 **Prompt Used**\n{image_prompt}\n\n### 🚀 **Next Steps**\n• **Download** the image by right-clicking and selecting 'Save Image As'\n• **Share** the image URL with others\n• **Generate more** images by describing what you'd like to see"
            else:
                # Fallback to text response
                assistant_message = result.get("choices", [{}])[0].get("message", {}).get("content", "Image generation completed.")
            
            # Add assistant message to session
            assistant_msg = {
                "role": "assistant",
                "content": assistant_message,
                "file_name": None,
                "is_image_response": True
            }
            chat_sessions[session_id].append(assistant_msg)
            
            # Increment usage for demo users
            increment_usage(user['email'], 100)  # Image generation costs more tokens
            
            return {"content": assistant_message, "file_name": file_name, "is_image_response": True}
            
    except Exception as e:
        error_message = f"🖼️ **Image Generation Error**\n\n### ❌ **Issue**\nFailed to generate image: {str(e)}\n\n### 💡 **Suggestions**\n• Try a different model that supports image generation\n• Check if your API key has image generation permissions\n• Ensure your prompt is clear and descriptive\n\n### 🚀 **Next Steps**\nTry asking for a text description instead or use a different model."
        
        assistant_msg = {
            "role": "assistant",
            "content": error_message,
            "file_name": None,
            "is_image_response": False
        }
        chat_sessions[session_id].append(assistant_msg)
        
        return {"error": error_message, "file_name": file_name}

@app.get("/", response_class=HTMLResponse)
async def chat_page(request: Request):
    """Main chat page - requires authentication"""
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    return templates.TemplateResponse("chat.html", {
        "request": request, 
        "version": VERSION,
        "user": user
    })

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    """Login page"""
    user = get_current_user(request)
    if user:
        return RedirectResponse(url="/", status_code=302)
    
    error = request.query_params.get("error")
    error_message = request.query_params.get("message", "")
    
    # Handle different error types
    if error == "access_denied":
        error_display = f"Access Denied: {error_message}"
    elif error == "invalid_state":
        error_display = "Invalid authentication state. Please try again."
    elif error == "no_code":
        error_display = "Authentication failed. Please try again."
    elif error == "auth_failed":
        error_display = "Authentication failed. Please try again."
    else:
        error_display = error if error else None
    
    return templates.TemplateResponse("login.html", {
        "request": request, 
        "version": VERSION,
        "error": error_display
    })

@app.get("/auth/google")
async def google_login(request: Request):
    """Initiate Google OAuth login"""
    return await login_google(request)

@app.get("/auth/callback")
async def google_callback(request: Request):
    """Handle Google OAuth callback"""
    return await auth_callback(request)

@app.get("/logout")
async def logout_user(request: Request):
    """Logout user"""
    return await logout(request)

@app.get("/api/models")
async def get_models(request: Request):
    """Get available models - requires authentication"""
    user = require_auth(request)
    
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {API_KEY}"
    }
    
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(f"{LITELLM_API_BASE}/models", headers=headers, timeout=10.0)
            response.raise_for_status()
            return response.json()
    except Exception as e:
        print(f"Error fetching models: {e}")
        return {"error": "Failed to fetch models", "data": []}

@app.post("/api/chat")
async def chat_completion(
    request: Request,
    message: str = Form(...),
    model: str = Form(...),
    session_id: str = Form(...),
    file_content: Optional[str] = Form(None),
    file_name: Optional[str] = Form(None),
    is_image_request: Optional[bool] = Form(False)
):
    """Chat completion endpoint - requires authentication"""
    user = require_auth(request)
    
    # Check demo limits for non-admin users
    can_proceed, error_message = check_demo_limits(user['email'])
    if not can_proceed:
        return {"error": error_message, "file_name": file_name}
    
    # Initialize session if it doesn't exist
    if session_id not in chat_sessions:
        chat_sessions[session_id] = []
    
    # Add user message to session
    user_message = {
        "role": "user",
        "content": message,
        "file_name": file_name if file_content else None,
        "is_image_request": is_image_request
    }
    chat_sessions[session_id].append(user_message)
    
    # Check if this is an image generation request
    should_generate_image = is_image_request or (is_image_generation_model(model) and not file_content)
    
    print(f"🔍 Debug: is_image_request={is_image_request}, model={model}, is_image_model={is_image_generation_model(model)}, has_file={bool(file_content)}, should_generate_image={should_generate_image}")
    
    if should_generate_image:
        print(f"🔍 Debug: Using image generation mode")
        return await handle_image_generation(request, message, model, session_id, file_content, file_name)
    else:
        print(f"🔍 Debug: Using regular chat completion mode")
    
    # Prepare the message for LiteLLM
    user_content = message
    file_info = ""
    max_file_chars = 4000
    
    # Create cache key for this request
    cache_key = f"{model}:{user_content}:{file_name or 'no_file'}"
    
    # Check cache first for faster responses
    if cache_key in response_cache:
        cached_response = response_cache[cache_key]
        assistant_msg = {
            "role": "assistant",
            "content": cached_response,
            "file_name": None
        }
        chat_sessions[session_id].append(assistant_msg)
        return {"content": cached_response, "file_name": file_name}
    
    if file_content and file_name:
        try:
            file_data = base64.b64decode(file_content)
            file_text = process_file_content(file_data, file_name)
            if file_text:
                if len(file_text) > max_file_chars:
                    file_text = file_text[:max_file_chars] + "\n... (truncated)"
                file_info = f"\n\n[Uploaded file: {file_name}]\n---\n{file_text}\n---"
            else:
                file_info = f"\n\n[Uploaded file: {file_name} (binary or non-text)]"
        except Exception as e:
            file_info = f"\n\n[Uploaded file: {file_name} (could not decode)]"
        user_content = f"{message}{file_info}"
    
    payload = {
        "model": model,
        "messages": [
            {
                "role": "system", 
                "content": """You are a helpful AI assistant with expertise in analyzing data, files, and providing insights. 

🚀 **CRITICAL FORMATTING REQUIREMENTS - ALWAYS FOLLOW THESE:**

1. **ALWAYS start responses with relevant emojis** (🎯, 📊, 🔍, 💡, etc.)
2. **ALWAYS use clear section headers** with ### or ** for structure
3. **ALWAYS include bullet points** (• or -) for lists and key points
4. **ALWAYS use numbered lists** (1., 2., 3.) for steps or sequences
5. **ALWAYS format code blocks** with proper syntax highlighting
6. **ALWAYS use bold text** (**text**) for emphasis on important terms
7. **ALWAYS include actionable recommendations** when applicable
8. **ALWAYS use tables or structured formats** for data presentation
9. **ALWAYS end with a call-to-action** or next steps
10. **NEVER return plain text** - always use rich formatting

📋 **RESPONSE STRUCTURE:**
- Start with emoji + bold title: 🎯 **Your Title Here**
- Use ### for main sections: ### 🔍 **Section Name**
- Use bullet points for lists: • **Key Point**
- Use numbered lists for steps: 1. **Step One**
- Use code blocks: ```python\ncode here\n```
- End with: ### 🚀 **Next Steps** + 💡 call-to-action

Your responses should be informative, visually appealing, and easy to understand."""
            },
            {"role": "user", "content": user_content}
        ],
        "stream": False,
        "max_tokens": 4096,
        "temperature": 0.5,
        "top_p": 0.9
    }
    
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {API_KEY}"
    }
    
    print(f"🔍 Debug: Sending request to {LITELLM_API_BASE}/chat/completions")
    print(f"🔍 Debug: Model: {model}")
    print(f"🔍 Debug: Headers: {headers}")
    print(f"🔍 Debug: Payload message length: {len(user_content)}")
    
    try:
        async with httpx.AsyncClient() as client:
            response = await client.post(
                f"{LITELLM_API_BASE}/chat/completions",
                json=payload,
                headers=headers,
                timeout=15.0
            )
            print(f"🔍 Debug: Response status: {response.status_code}")
            print(f"🔍 Debug: Response headers: {dict(response.headers)}")
            result = response.json()
            print(f"🔍 Debug: Response body: {result}")
            
            if "error" in result:
                error_msg = result["error"]
                if "NotFoundError" in error_msg:
                    # Extract model name from error message
                    import re
                    model_match = re.search(r"'([^']+)'", error_msg)
                    if model_match:
                        model_name = model_match.group(1)
                        return {"error": f"Model '{model_name}' not found. Please select a different model from the dropdown.", "file_name": file_name}
                    else:
                        return {"error": "Model not found. Please select a different model from the dropdown.", "file_name": file_name}
                else:
                    return {"error": error_msg, "file_name": file_name}
            
            if "choices" not in result or not result["choices"]:
                return {"error": "Invalid response format from API", "file_name": file_name}
            
            assistant_message = result.get("choices", [{}])[0].get("message", {}).get("content", "No response received")
            if not assistant_message or assistant_message == "No response received":
                return {"error": "No response content received from API", "file_name": file_name}
            
            # Add assistant message to session
            assistant_msg = {
                "role": "assistant",
                "content": assistant_message,
                "file_name": None
            }
            chat_sessions[session_id].append(assistant_msg)
            
            # Cache the response for future use
            response_cache[cache_key] = assistant_message
            
            # Limit cache size to prevent memory issues
            if len(response_cache) > 100:
                # Remove oldest entries
                oldest_keys = list(response_cache.keys())[:20]
                for key in oldest_keys:
                    del response_cache[key]
            
            # Increment usage for demo users
            estimated_tokens = len(assistant_message.split()) + len(user_content.split())  # Rough estimation
            increment_usage(user['email'], estimated_tokens)
            
            return {"content": assistant_message, "file_name": file_name}
            
    except httpx.HTTPStatusError as e:
        print(f"🔍 Debug: HTTP Status Error: {e.response.status_code} - {e.response.text}")
        if e.response.status_code == 404:
            return {"error": f"Model '{model}' not found. Please select a different model from the dropdown.", "file_name": file_name}
        elif e.response.status_code == 401:
            return {"error": "Authentication failed. Please check your API key.", "file_name": file_name}
        elif e.response.status_code == 403:
            return {"error": "Access forbidden. Please check your API key and permissions.", "file_name": file_name}
        else:
            return {"error": f"API error ({e.response.status_code}): {e.response.text}", "file_name": file_name}
    except httpx.TimeoutException:
        return {"error": "Request timed out. Please try again.", "file_name": file_name}
    except httpx.ConnectError:
        return {"error": "Could not connect to the API server. Please check your internet connection.", "file_name": file_name}
    except json.JSONDecodeError as e:
        print(f"🔍 Debug: JSON Decode Error: {e}")
        return {"error": "Invalid JSON response from API", "file_name": file_name}
    except Exception as e:
        print(f"🔍 Debug: Unexpected Error: {type(e).__name__}: {str(e)}")
        return {"error": f"Unexpected error: {type(e).__name__}: {str(e)}", "file_name": file_name}

@app.get("/api/chat-history/{session_id}")
async def get_chat_history(session_id: str, request: Request):
    """Get chat history for a session - requires authentication"""
    user = require_auth(request)
    
    if session_id in chat_sessions:
        return {"messages": chat_sessions[session_id]}
    else:
        return {"messages": []}

@app.post("/api/new-session")
async def create_new_session(request: Request):
    """Create a new chat session - requires authentication"""
    user = require_auth(request)
    
    session_id = str(uuid.uuid4())
    chat_sessions[session_id] = []
    return {"session_id": session_id}

@app.get("/api/usage")
async def get_usage(request: Request):
    """Get current usage for the authenticated user"""
    user = require_auth(request)
    
    if user['is_admin']:
        return {
            "is_admin": True,
            "usage": {
                "request_count": "Unlimited",
                "token_count": "Unlimited",
                "limit_reached": False
            }
        }
    else:
        usage = user['usage']
        return {
            "is_admin": False,
            "usage": {
                "request_count": usage['request_count'],
                "token_count": usage['token_count'],
                "request_limit": int(os.getenv('DEMO_REQUEST_LIMIT', '2')),
                "token_limit": int(os.getenv('DEMO_TOKEN_LIMIT', '100')),
                "limit_reached": usage['request_count'] >= int(os.getenv('DEMO_REQUEST_LIMIT', '2'))
            }
        }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000) 